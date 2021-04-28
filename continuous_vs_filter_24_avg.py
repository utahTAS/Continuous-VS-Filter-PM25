# -*- coding: utf-8 -*-
"""
Created on Thu Apr 15 15:13:26 2021

@author: bcubrich
"""


#%%

import math
import pyodbc
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

from tkinter import Tk
from tkinter.filedialog import askopenfilename
import os
from tkinter import *
import pyodbc 
import datetime as dt  
import time
import sys
from bs4 import BeautifulSoup
#from tkinter.filedialog import asksavefilename
from tkinter.filedialog import askdirectory
    # Import smtplib for the actual sending function
import smtplib

# Import the email modules we'll need
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
import seaborn as sns
import json
import requests
import scipy

from operator import sub

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
import shutil

import datetime as dt

driver_names = pyodbc.drivers()  

#%%
start_date_dt=pd.to_datetime('12-31-2020')
end_date_dt=pd.to_datetime('03-01-2021')
av_start_date=str(start_date_dt)[:-3]
av_end_date=str(end_date_dt)[:-3]

sites=r'U:/PLAN/BCUBRICH/Python/Parameter Reader/'\
r'PARAMETERS.xls'

sites_df=pd.read_excel(sites, converters={'SITE NAME':str,'State Code':str,
                                          'County Code':str, 'Site Code':str,
                                          'Paramter':str, 'Analyt':str, 
                                          'Method':str, 'Unit':str}) # load data
sites_df['Analyt']=sites_df['Analyt'].str.strip('()') #strip parentheses from 


site_dict={x:[y,z] for x,y,z in zip(sites_df['Site Symbol'],sites_df['County Code'],sites_df['Site Code'])}

#%%
def av_query(start_date,end_date,param):
    
    driver_names = pyodbc.drivers()           #get drivers that will work in the current env.
#    print(driver_names)
    #https://www.microsoft.com/en-us/download/details.aspx?id=56567, need to get driver from here
#    db='AVData.Reporting.CalibrationDataFull' #this is where the data comes from
    #Need to login to the server with these
    username='gis'
    sql_password='Axd!35jkl'
    
    cnxn = pyodbc.connect(r'DRIVER={'+driver_names[-1]+'};'
                            r'SERVER=168.178.43.251;'
                            r'DATABASE=AVData;'
                            r'UID='+username+';PWD='+sql_password+';'
                        r'timeout=10')
    
    cursor = cnxn.cursor()
    row_names=[]
    df_tab=pd.DataFrame()
    
    get_table_names=False


    query="""SELECT 'RD' AS TransactionType, 'I' AS ActionIndicator,  
        1113 As PerformingAgencyCode,
        AqsStateCode AS State, 
        AqsSiteCode AS Site,
        AqsParameterCode AS ParameterCode,
        ParameterName AS ParameterName,
        AqsParameterOccuranceCode AS POC,
        Date,
        RawValue,
        NullCode,
        '1' As AssessmentNumber,
        ParameterAqsMethodCode AS MethodCode,
        AqsUnitCode AS ReportingUnit,
        SiteAbbreviation
        FROM AVData.Reporting.ReadingAverageDataFull
        
        
        
        WHERE SiteAbbreviation IN ('BR', 'SM', 'P2', 'BV', 'RS', 'ES', 'EN', 'CV', 'HW', 'RP', 'H3', 'ED', 'V4', 'LN', 'SF', 'O2', 'HV', 'HC', 'AI', 'SA', 'NR', 'MA', 'UT')
        AND Date > '{}'
        AND Date < '{}'
        AND [IntervalName] =  '001h'
        AND AqsParameterCode ='{}'
        
        AND ParameterEnabled = 1
        ORDER BY AqsStateCode, 
        AqsCountyTribalCode,
        AqsSiteCode,
        AqsParameterCode,
        Date""".format(start_date,end_date,param)
        
    df = pd.read_sql_query(query, cnxn)
    
    cnxn.close()
    
    return df


def get_AQS_data(param='88101',
                 bdate=pd.datetime.now()-pd.Timedelta('{} days'.format(int(30))),
                 edate=pd.datetime.now(),
                 state='49',
                 county='035',
                 site='3006'):
    global request
    df=pd.DataFrame()
    if  edate.year != bdate.year:
        for year in range(bdate.year,edate.year+1):
            if year==range(bdate.year,edate.year+1)[0]:
                bdate_=pd.to_datetime(bdate)
                edate_=pd.to_datetime('12-31-{} 23:59'.format(year))
    #            print(bdate)
                request=get_AQS_url(email='bcubrich@utah.gov',
                 key='ecruhawk94',
                 bdate=bdate_.strftime('%Y%m%d'), 
                 edate=edate_.strftime('%Y%m%d'), 
                 param=param, 
                 state=state, 
                 county=county, 
                 site=site) 
                response = requests.get(request)
                aqs_data = json.loads(response.text).get('Data')
                df = df.append(pd.DataFrame.from_dict(aqs_data,orient='columns'))
                
            elif year==range(bdate.year,edate.year+1)[-1]:
                bdate_=pd.to_datetime('01-01-{} 00:00'.format(year))
                edate_=pd.to_datetime(edate)
    #            print(bdate)
                request=get_AQS_url(email='bcubrich@utah.gov',
                 key='ecruhawk94',
                 bdate=bdate_.strftime('%Y%m%d'), 
                 edate=edate_.strftime('%Y%m%d'), 
                 param=param, 
                 state=state, 
                 county=county, 
                 site=site) 
                response = requests.get(request)
                aqs_data = json.loads(response.text).get('Data')
                df = df.append(pd.DataFrame.from_dict(aqs_data,orient='columns'))
            
            else:
                bdate_=pd.to_datetime('01-01-{} 00:00'.format(year))
                edate_=pd.to_datetime('12-31-{} 23:59'.format(year))
    #            print(bdate)
                request=get_AQS_url(email='bcubrich@utah.gov',
                 key='ecruhawk94',
                 bdate=bdate_.strftime('%Y%m%d'), 
                 edate=edate_.strftime('%Y%m%d'), 
                 param=param, 
                 state=state, 
                 county=county, 
                 site=site) 
                response = requests.get(request)
                aqs_data = json.loads(response.text).get('Data')
                df = df.append(pd.DataFrame.from_dict(aqs_data,orient='columns'))
                
    else:        
    
        request=get_AQS_url(email='bcubrich@utah.gov',
                     key='ecruhawk94',
                     bdate=bdate.strftime('%Y%m%d'), 
                     edate=edate.strftime('%Y%m%d'), 
                     param=param, 
                     state=state, 
                     county=county, 
                     site=site) 

        response = requests.get(request)
    #    print(request)
        aqs_data = json.loads(response.text).get('Data')
    

        df = pd.DataFrame.from_dict(aqs_data,orient='columns')
#    print(df)
#    print(df.columns)
    try:
        df['dt']=pd.to_datetime(df['date_local']+' '+df['time_local'])
    except KeyError:
        do_nothin=1
#    df.reset_index(level=0, inplace=True)
    if len(df)>=1:    
        df=df[(df['dt']<edate) & (df['dt']>bdate)]
    
    return df

station_sym_dict={'490110004':'BV', '490450004':'ED','490170006':'ES','490351007':'MA',
          '490030003':'BR','490571003':'HV','490570002':'O2', '490050007':'SM',
          '490494001':'LN', '490354002':'NR','490130002':'RS', '490495010':'SF', '490471004':'V4',
          '490353013':'H3','490353006':'HW','490071003':'P2','490116001':'AI','490456001':'490456001','490353005':'SA',
          '490352005':'CV','490210005':'EN','490530007':'HC','490353010':'RP','490353015':'UT','490353014':'LP','490353016':'IP'}

#%%
def get_AQS_url(email, key, bdate, edate, param, state, county, site):
    url = "https://aqs.epa.gov/data/api/sampleData/bySite?email={}&key={}&bdate={}&edate={}&param={}&state={}&county={}&site={}".format(email, key, bdate, edate, param, state, county, site)
    print(url)
    return url

aqs_df=pd.DataFrame()

site_list=sites_df['Site Symbol'].unique()
for site_ in site_list:
    site=site_dict.get(site_)
    temp_df=get_AQS_data(param='88101',bdate=start_date_dt,edate=end_date_dt,site=site[1],county=site[0])
    # temp_df=temp_df[temp_df['sample_duration']=='24 HOUR']
    aqs_df=aqs_df.append(temp_df)
#%%
aqs_df['SiteCode']=aqs_df['state_code']+aqs_df['county_code']+aqs_df['site_number']
aqs_df['SiteSym']=aqs_df['SiteCode'].map(station_sym_dict)
# aqs_df=aqs_df[aqs_df['sample_frequency']!='EVERY 12TH DAY']
# aqs_df=aqs_df[aqs_df['sample_frequency']!='EVERY 6TH DAY']

filter_df=aqs_df[aqs_df['sample_duration']=='24 HOUR']
continuous_df=aqs_df[aqs_df['sample_duration']=='1 HOUR']


aqs_df_24_hr_cont=continuous_df.groupby(['date_local','poc','SiteSym']).agg(np.nanmean)

aqs_df_24_hr_cont=aqs_df_24_hr_cont.rename({'sample_measurement':'Continuous MC'})
aqs_df_24_hr_cont=aqs_df_24_hr_cont['sample_measurement']
aqs_df_24_hr_cont=aqs_df_24_hr_cont.reset_index()


#%%


# av_df=av_query(start_date,end_date,'88101')

# av_df['Date Group']=av_df['Date'].dt.strftime('%Y%m%d')

# av_df_24_hr=av_df.groupby(['Date Group','ParameterName','POC','SiteAbbreviation']).mean()

save_location='U:/PLAN/BCUBRICH/Monthly QA Reports/2021/'

# save_location+=start_date_dt.strftime('%B')

save_location+='/{}.xlsx'.format(dt.datetime.now().strftime('%Y%m%d'))

import openpyxl
from openpyxl import load_workbook
wb = openpyxl.Workbook()
wb.save(save_location)
wb.close()

aqs_df.columns
output_dfs={}
for site in aqs_df['SiteSym'].unique():
    # print(site)
    book=load_workbook(save_location)
    writer = pd.ExcelWriter(save_location, engine = 'openpyxl')
    writer.book=book
    # output_df=1
    temp_filter_df1=filter_df[filter_df['SiteSym']==site]
    temp_filter_df1=temp_filter_df1[['date_local','SiteSym','poc','sample_measurement']]
    
    
    if len(temp_filter_df1)>0:
        output_df_filter=pd.DataFrame({'date_local':temp_filter_df1['date_local'].unique(),'SiteSym':site})
        old_poc=temp_filter_df1['poc'].unique()[0]
        
        
    if len(temp_filter_df1['poc'].unique())>1:
        temp_filter_df1=temp_filter_df1.rename({'sample_measurement':'filter'},axis='columns')
        for poc in temp_filter_df1['poc'].unique():
            
            temp_filter_df2=temp_filter_df1[temp_filter_df1['poc']==poc]
            output_df_filter=output_df_filter.merge(temp_filter_df2,
                                                    on=['date_local','SiteSym'],
                                                    suffixes=['','_poc'+str(poc)]
                                                    ,how='outer')
            # output_df_filter=output_df_filter.rename({'filter':'filter_poc'+str(old_poc)},axis='columns')
            old_poc=poc
                

            
    else:
        temp_filter_df1=temp_filter_df1.rename({'sample_measurement':'filter_poc'+str(old_poc)},axis='columns')
        output_df_filter=temp_filter_df1
        output_df_filter=output_df_filter.drop(['poc'],axis='columns')
    
            
        # print(output_df_filter)    
        
    temp_continuous_df1=aqs_df_24_hr_cont[aqs_df_24_hr_cont['SiteSym']==site]
    temp_continuous_df1=temp_continuous_df1[['date_local','SiteSym','poc','sample_measurement']]
    
    
    if len(temp_continuous_df1)>0:
        output_df_continuous=pd.DataFrame({'date_local':temp_continuous_df1['date_local'].unique(),'SiteSym':site})
        old_poc=temp_continuous_df1['poc'].unique()[0]
        
        
        print(len(temp_continuous_df1['poc'].value_counts())  ) 
    
    if len(temp_continuous_df1['poc'].unique())>1:
        temp_continuous_df1=temp_continuous_df1.rename({'sample_measurement':'continuous'},axis='columns')
        for poc in temp_continuous_df1['poc'].unique():
            
            temp_continuous_df2=temp_continuous_df1[temp_continuous_df1['poc']==poc]
            output_df_continuous=output_df_continuous.merge(temp_continuous_df2,
                                                            on=['date_local','SiteSym'],
                                                            suffixes=['','_cont_poc'+str(poc)]
                                                            ,how='outer')
            
            # output_df_continuous=output_df_continuous.rename({'continuous':'continuous_poc'+str(old_poc)},axis='columns')
            old_poc=poc
        
        # output_df_continuous=output_df_continuous.drop(['poc_x','poc_y'])
        
    else:
        temp_continuous_df1=temp_continuous_df1.rename({'sample_measurement':'continuous_poc'+str(old_poc)},axis='columns')
        output_df_continuous=temp_continuous_df1
        output_df_continuous=output_df_continuous.drop(['poc'],axis='columns')
            
           
           
    
    if len(output_df_filter)>0 and len(output_df_continuous)>0:
        output_df=output_df_continuous.merge(output_df_filter,on=['date_local','SiteSym'],how='outer')
    elif len(output_df_filter)>0:
        output_df=output_df_filter
    elif len(output_df_continuous)>0:
        output_df=output_df_continuous
   
    
    output_df = output_df.loc[:, ~output_df.columns.str.startswith('poc')]
    
    output_df['Date']=pd.to_datetime(output_df['date_local'])
    output_df['Date']=output_df['Date'].dt.strftime('%m/%d/%Y')
    
    
    output_df=output_df.drop(['SiteSym','date_local'],axis='columns')
    cols=output_df.columns.to_list()
    cols = cols[-1:] + cols[:-1]
    output_df=output_df[cols]
    output_df.to_excel(writer, sheet_name=site,index=False)
            
        
    output_dfs[site]=output_df
    
    writer.save()
    writer.close()
        
        # print(output_df)